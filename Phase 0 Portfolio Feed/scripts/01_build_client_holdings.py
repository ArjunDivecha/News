#!/usr/bin/env python3
"""
=============================================================================
SCRIPT NAME: 01_build_client_holdings.py
=============================================================================

DESCRIPTION:
    Build a single normalized holdings file (Client.xlsx) from Schwab and
    Interactive Brokers (IBKR) position exports. Optionally runs the source
    broker fetch scripts (Schwab All Accounts Data.py and IBKR.py) as
    subprocesses to refresh the export data before loading. Normalizes each
    broker's position data into a unified schema, aggregates positions by
    ticker symbol across all accounts (summing market value, quantity, and
    P&L; computing a quantity-weighted average price), and writes the result
    to a fixed-format Excel workbook with 5 columns: Symbol, Market Value,
    Average Price, Long Quantity, Long Open Profit/Loss. Supports preflight
    checks (Schwab OAuth token freshness, IBKR API port readiness with
    optional auto-launch of TWS/IB Gateway) and optional timestamped backup
    of any prior output file before overwriting.

INPUT FILES:
    /Users/arjundivecha/Dropbox/AAA Backup/Master Valuation/Schwab All Accounts Data.xlsx
        Schwab position export (produced by Schwab All Accounts Data.py).
        Columns: Account, Name, Value, Bloom, Shares, AvgCost, USD P&L.
    /Users/arjundivecha/Dropbox/AAA Backup/Master Valuation/IBKR Account Data.xlsx
        IBKR position export (produced by IBKR.py).
        Columns: Account, Name, Value, Shares, AvgCost, Unrealized P&L, Currency.
    /Users/arjundivecha/Dropbox/AAA Backup/A Working/News/Client.xlsx
        Prior output file (if it exists) -- read in to clear and rewrite the
        first sheet rather than appending.

OUTPUT FILES:
    /Users/arjundivecha/Dropbox/AAA Backup/A Working/News/Client.xlsx
        Normalized aggregated client holdings: Symbol, Market Value, Average
        Price, Long Quantity, Long Open Profit/Loss.
    /Users/arjundivecha/Dropbox/AAA Backup/A Working/News/Client.backup_YYYYMMDD_HHMMSS.xlsx
        Optional timestamped backup of the prior Client.xlsx created before
        overwriting (controlled by --no-backup flag).

VERSION: 1.0
LAST UPDATED: 2026-06-05
AUTHOR: Arjun Divecha

DEPENDENCIES:
    - pandas
    - openpyxl

USAGE:
    python 01_build_client_holdings.py
    python 01_build_client_holdings.py --no-refresh-schwab --no-refresh-ibkr
    python 01_build_client_holdings.py --skip-schwab --include-cash

NOTES:
    - Source broker scripts are run as subprocesses by default to refresh
      export data before loading. Use --no-refresh-* flags to skip.
    - IBKR preflight probes port 7496 (TWS live trading port) and can
      auto-launch Trader Workstation or IB Gateway via macOS open.
    - Cash positions (rows whose name/symbol indicates cash) are excluded
      by default; use --include-cash to keep them.
    - Schwab OAuth token DB is checked for existence/mtime only; actual
      refresh/auth is handled by the Schwab source script.
============================================================================="""

from __future__ import annotations

import argparse
import importlib.util
import numbers
import shutil
import socket
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook


DEFAULT_SCHWAB_SCRIPT = Path(
    "/Users/arjundivecha/Dropbox/AAA Backup/Master Valuation/Schwab All Accounts Data.py"
)
DEFAULT_IBKR_SCRIPT = Path(
    "/Users/arjundivecha/Dropbox/AAA Backup/Master Valuation/IBKR.py"
)
DEFAULT_SCHWAB_XLSX = Path(
    "/Users/arjundivecha/Dropbox/AAA Backup/Master Valuation/Schwab All Accounts Data.xlsx"
)
DEFAULT_IBKR_XLSX = Path(
    "/Users/arjundivecha/Dropbox/AAA Backup/Master Valuation/IBKR Account Data.xlsx"
)
DEFAULT_OUTPUT = Path(
    "/Users/arjundivecha/Dropbox/AAA Backup/A Working/News/Client.xlsx"
)

CLIENT_HEADERS = [
    "Symbol",
    "Market Value",
    "Average Price",
    "Long Quantity",
    "Long Open Profit/Loss",
]


def run_source_script(script_path: Path, label: str, python_executable: str) -> None:
    """Run broker fetch script in-place so its relative output path stays stable."""
    if not script_path.exists():
        raise FileNotFoundError(f"{label} script not found: {script_path}")

    print(f"\n[{label}] Running source script:")
    print(f"  {script_path}")
    print(f"  Python: {python_executable}")

    try:
        result = subprocess.run(
            [python_executable, str(script_path)],
            cwd=str(script_path.parent),
            capture_output=False,
            text=True,
            check=False,
        )
    except FileNotFoundError as exc:
        raise RuntimeError(
            f"{label} script Python executable not found: {python_executable}"
        ) from exc

    if result.returncode != 0:
        raise RuntimeError(f"{label} script failed with exit code {result.returncode}")


def require_module(
    module_name: str,
    source_label: str,
    pip_hint: str,
    python_executable: str,
) -> None:
    """Fail fast with a clear message when a source dependency is missing."""
    if python_executable == sys.executable:
        module_found = importlib.util.find_spec(module_name) is not None
    else:
        try:
            probe = subprocess.run(
                [python_executable, "-c", f"import {module_name}"],
                capture_output=True,
                text=True,
                check=False,
            )
        except FileNotFoundError as exc:
            raise RuntimeError(
                f"Source Python executable not found: {python_executable}"
            ) from exc
        module_found = probe.returncode == 0

    if module_found:
        return

    raise RuntimeError(
        f"{source_label} refresh requires Python package '{module_name}' in interpreter "
        f"'{python_executable}'. Install with: {pip_hint} | or run with source refresh disabled."
    )


def is_port_open(host: str, port: int, timeout: float = 1.0) -> bool:
    """Return True when a TCP port is reachable."""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(timeout)
        return sock.connect_ex((host, port)) == 0


def find_tws_app_path(explicit_path: str | None = None) -> Path | None:
    """Resolve a likely Trader Workstation / IB Gateway app path on macOS."""
    candidates = []
    if explicit_path:
        candidates.append(Path(explicit_path).expanduser())

    home_apps = Path.home() / "Applications"
    candidates.extend(
        [
            home_apps / "Trader Workstation" / "Trader Workstation.app",
            home_apps / "Trader Workstation.app",
            Path("/Applications/Trader Workstation.app"),
            Path("/Applications/Trader Workstation/Trader Workstation.app"),
            home_apps / "IB Gateway" / "IB Gateway.app",
            Path("/Applications/IB Gateway.app"),
            Path("/Applications/IB Gateway/IB Gateway.app"),
        ]
    )

    for app in candidates:
        if app.exists():
            return app

    # Fallback scan for non-standard install names.
    patterns = [
        "*Trader*Workstation*.app",
        "*TWS*.app",
        "*IB*Gateway*.app",
        "*Interactive*Brokers*.app",
    ]
    for root in [Path("/Applications"), home_apps]:
        if not root.exists():
            continue
        for pattern in patterns:
            matches = sorted(root.glob(pattern))
            if matches:
                return matches[0]

    return None


def ensure_ibkr_ready(port: int, tws_app: str | None, wait_seconds: int) -> None:
    """
    Ensure IBKR API endpoint is reachable; auto-start TWS/IB Gateway if needed.
    """
    if is_port_open("127.0.0.1", port):
        print(f"  IBKR preflight: API port 127.0.0.1:{port} is already open")
        return

    app_path = find_tws_app_path(tws_app)
    if app_path is None:
        raise RuntimeError(
            "IBKR preflight failed: could not find Trader Workstation/IB Gateway app. "
            "Set --tws-app explicitly."
        )

    print(f"  IBKR preflight: API port {port} not open, launching app...")
    print(f"    App: {app_path}")
    subprocess.run(["open", str(app_path)], check=False)

    deadline = time.time() + wait_seconds
    while time.time() < deadline:
        if is_port_open("127.0.0.1", port):
            print(f"  IBKR preflight: API port opened after app launch")
            return
        time.sleep(2)

    raise RuntimeError(
        f"IBKR preflight failed: API port 127.0.0.1:{port} still closed after "
        f"{wait_seconds}s. Log in to TWS/IB Gateway and enable API access."
    )


def check_schwab_token() -> None:
    """Basic Schwab token preflight check."""
    token_db = Path.home() / ".schwabdev" / "tokens.db"
    if not token_db.exists():
        print("  Schwab preflight: token DB not found; Schwab script should open auth flow.")
        return

    mtime = datetime.fromtimestamp(token_db.stat().st_mtime)
    age_hours = (datetime.now() - mtime).total_seconds() / 3600
    print(
        f"  Schwab preflight: token DB found (last updated {mtime:%Y-%m-%d %H:%M:%S}, "
        f"{age_hours:.1f}h ago)"
    )
    print("    If expired, Schwab script will open the auth window automatically.")


def load_excel(path: Path, label: str) -> pd.DataFrame:
    """Load source workbook into dataframe."""
    if not path.exists():
        raise FileNotFoundError(f"{label} file not found: {path}")

    df = pd.read_excel(path)
    print(f"  Loaded {label}: {len(df):,} rows")
    return df


def parse_schwab_symbol(bloom: object, name: object) -> str:
    """Extract ticker from Bloomberg-formatted symbol (e.g., 'EWI Equity')."""
    bloom_text = str(bloom).strip() if bloom is not None else ""
    if bloom_text:
        ticker = bloom_text.split()[0].strip().upper()
        if ticker and ticker != "N/A":
            return ticker

    name_text = str(name).strip() if name is not None else ""
    if name_text.upper() == "CASH":
        return "CASH"
    return ""


def is_cash_position(symbol: str, name: str, currency: str | None = None) -> bool:
    """Identify rows that represent cash balances instead of tradeable holdings."""
    # Handle NaN/float values by converting to string
    # pd.isna() handles None, NaN, and NaT
    symbol_str = "" if (symbol is None or pd.isna(symbol)) else str(symbol)
    name_str = "" if (name is None or pd.isna(name)) else str(name)
    currency_str = "" if (currency is None or pd.isna(currency)) else str(currency)
    
    symbol_upper = symbol_str.strip().upper()
    name_upper = name_str.strip().upper()
    currency_upper = currency_str.strip().upper()

    if name_upper in {"CASH", "BASE CASH"}:
        return True
    if symbol_upper in {"CASH", "SNAXX", "BASE_CASH"}:
        return True
    if "CASH" in name_upper and currency_upper == "BASE":
        return True
    return False


def _require_columns(df: pd.DataFrame, cols: Iterable[str], label: str) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise ValueError(f"{label} missing required columns: {missing}")


def normalize_schwab(df: pd.DataFrame, include_cash: bool) -> pd.DataFrame:
    """Normalize Schwab export to unified schema."""
    _require_columns(
        df,
        ["Account", "Name", "Value", "Bloom", "Shares", "AvgCost", "USD P&L"],
        "Schwab",
    )

    out = pd.DataFrame(
        {
            "source": "Schwab",
            "account": df["Account"].astype(str).str.strip(),
            "name": df["Name"].astype(str).str.strip(),
            "symbol": df.apply(
                lambda r: parse_schwab_symbol(r.get("Bloom"), r.get("Name")), axis=1
            ),
            "market_value": pd.to_numeric(df["Value"], errors="coerce"),
            "quantity": pd.to_numeric(df["Shares"], errors="coerce"),
            "avg_price": pd.to_numeric(df["AvgCost"], errors="coerce"),
            "open_pnl": pd.to_numeric(df["USD P&L"], errors="coerce"),
            "currency": "USD",
        }
    )

    out = out[out["symbol"].notna() & (out["symbol"].str.len() > 0)]
    if not include_cash:
        out = out[
            ~out.apply(
                lambda r: is_cash_position(r["symbol"], r["name"], r["currency"]), axis=1
            )
        ]
    return out.reset_index(drop=True)


def normalize_ibkr(df: pd.DataFrame, include_cash: bool) -> pd.DataFrame:
    """Normalize IBKR export to unified schema."""
    _require_columns(
        df,
        ["Account", "Name", "Value", "Shares", "AvgCost", "Unrealized P&L", "Currency"],
        "IBKR",
    )

    out = pd.DataFrame(
        {
            "source": "IBKR",
            "account": df["Account"].astype(str).str.strip(),
            "name": df["Name"].astype(str).str.strip(),
            "symbol": df["Name"].astype(str).str.strip().str.upper(),
            "market_value": pd.to_numeric(df["Value"], errors="coerce"),
            "quantity": pd.to_numeric(df["Shares"], errors="coerce"),
            "avg_price": pd.to_numeric(df["AvgCost"], errors="coerce"),
            "open_pnl": pd.to_numeric(df["Unrealized P&L"], errors="coerce"),
            "currency": df["Currency"].astype(str).str.strip(),
        }
    )

    out = out[out["symbol"].notna() & (out["symbol"].str.len() > 0)]
    if not include_cash:
        out = out[
            ~out.apply(
                lambda r: is_cash_position(r["symbol"], r["name"], r["currency"]), axis=1
            )
        ]
    return out.reset_index(drop=True)


def aggregate_positions(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate positions by symbol across all source accounts."""
    if df.empty:
        return pd.DataFrame(
            columns=["symbol", "market_value", "avg_price", "quantity", "open_pnl"]
        )

    work = df.copy()
    work["abs_quantity"] = work["quantity"].abs()
    work["weighted_avg_cost"] = work["avg_price"] * work["abs_quantity"]

    grouped = (
        work.groupby("symbol", as_index=False)
        .agg(
            market_value=("market_value", "sum"),
            quantity=("quantity", "sum"),
            open_pnl=("open_pnl", "sum"),
            abs_quantity=("abs_quantity", "sum"),
            weighted_avg_cost=("weighted_avg_cost", "sum"),
        )
        .fillna(0.0)
    )

    grouped["avg_price"] = grouped["weighted_avg_cost"] / grouped["abs_quantity"]
    grouped.loc[grouped["abs_quantity"] == 0, "avg_price"] = pd.NA
    grouped = grouped[
        ["symbol", "market_value", "avg_price", "quantity", "open_pnl"]
    ].sort_values(by="market_value", key=lambda s: s.abs(), ascending=False)

    return grouped.reset_index(drop=True)


def to_client_schema(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns to Phase 2 expected portfolio-file shape."""
    return df.rename(
        columns={
            "symbol": "Symbol",
            "market_value": "Market Value",
            "avg_price": "Average Price",
            "quantity": "Long Quantity",
            "open_pnl": "Long Open Profit/Loss",
        }
    )[CLIENT_HEADERS]


def _to_cell_value(value: object) -> object:
    if pd.isna(value):
        return None
    return float(value) if isinstance(value, numbers.Real) else value


def write_client_workbook(client_df: pd.DataFrame, output_path: Path, backup: bool) -> Path | None:
    """Write normalized holdings into Client.xlsx (first sheet)."""
    backup_path = None

    if output_path.exists():
        if backup:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = output_path.with_name(f"{output_path.stem}.backup_{ts}{output_path.suffix}")
            shutil.copy2(output_path, backup_path)

        wb = load_workbook(output_path)
        ws = wb[wb.sheetnames[0]]
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    ws.append(CLIENT_HEADERS)
    for _, row in client_df.iterrows():
        ws.append([_to_cell_value(row[col]) for col in CLIENT_HEADERS])

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return backup_path


def print_source_summary(df: pd.DataFrame, label: str) -> None:
    if df.empty:
        print(f"  {label}: 0 rows")
        return

    gross_mv = df["market_value"].abs().sum()
    net_mv = df["market_value"].sum()
    print(f"  {label}: {len(df):,} rows | Gross MV {gross_mv:,.2f} | Net MV {net_mv:,.2f}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Phase 0: build Client.xlsx from Schwab + IBKR exports"
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Output Client.xlsx path",
    )
    parser.add_argument(
        "--schwab-script",
        default=str(DEFAULT_SCHWAB_SCRIPT),
        help="Path to Schwab source script",
    )
    parser.add_argument(
        "--ibkr-script",
        default=str(DEFAULT_IBKR_SCRIPT),
        help="Path to IBKR source script",
    )
    parser.add_argument(
        "--schwab-file",
        default=str(DEFAULT_SCHWAB_XLSX),
        help="Path to Schwab export workbook",
    )
    parser.add_argument(
        "--ibkr-file",
        default=str(DEFAULT_IBKR_XLSX),
        help="Path to IBKR export workbook",
    )
    parser.add_argument(
        "--source-python",
        default=sys.executable,
        help="Python executable used for source scripts unless overridden",
    )
    parser.add_argument(
        "--schwab-python",
        default=None,
        help="Optional Python executable override for Schwab source script",
    )
    parser.add_argument(
        "--ibkr-python",
        default=None,
        help="Optional Python executable override for IBKR source script",
    )
    parser.add_argument(
        "--refresh-schwab",
        dest="refresh_schwab",
        action="store_true",
        default=True,
        help="Run Schwab source script before reading workbook (default: on)",
    )
    parser.add_argument(
        "--no-refresh-schwab",
        dest="refresh_schwab",
        action="store_false",
        help="Skip running Schwab source script and use existing workbook",
    )
    parser.add_argument(
        "--refresh-ibkr",
        dest="refresh_ibkr",
        action="store_true",
        default=True,
        help="Run IBKR source script before reading workbook (default: on)",
    )
    parser.add_argument(
        "--no-refresh-ibkr",
        dest="refresh_ibkr",
        action="store_false",
        help="Skip running IBKR source script and use existing workbook",
    )
    # Backward-compatible aliases.
    parser.add_argument(
        "--run-schwab",
        dest="refresh_schwab",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--run-ibkr",
        dest="refresh_ibkr",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--skip-schwab",
        action="store_true",
        help="Exclude Schwab source from aggregation",
    )
    parser.add_argument(
        "--skip-ibkr",
        action="store_true",
        help="Exclude IBKR source from aggregation",
    )
    parser.add_argument(
        "--include-cash",
        action="store_true",
        help="Keep cash rows (default is to exclude them)",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Disable pre-write backup for existing output",
    )
    parser.add_argument(
        "--ibkr-port",
        type=int,
        default=7496,
        help="IBKR API port to probe for readiness (default: 7496)",
    )
    parser.add_argument(
        "--tws-app",
        default=None,
        help="Optional explicit path to Trader Workstation/IB Gateway .app",
    )
    parser.add_argument(
        "--tws-wait-seconds",
        type=int,
        default=90,
        help="Seconds to wait for IBKR API port after app launch",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    output_path = Path(args.output)
    schwab_script = Path(args.schwab_script)
    ibkr_script = Path(args.ibkr_script)
    schwab_file = Path(args.schwab_file)
    ibkr_file = Path(args.ibkr_file)
    source_python = args.source_python
    schwab_python = args.schwab_python or source_python
    ibkr_python = args.ibkr_python or source_python

    print("=" * 80)
    print("PHASE 0 - PORTFOLIO HOLDINGS AGGREGATION")
    print("=" * 80)
    print(f"Output: {output_path}")
    print(f"Include Cash: {args.include_cash}")
    print(f"Refresh Schwab: {args.refresh_schwab}")
    print(f"Refresh IBKR: {args.refresh_ibkr}")
    print(f"Source Python (default): {source_python}")
    if args.schwab_python:
        print(f"Schwab Python override: {args.schwab_python}")
    if args.ibkr_python:
        print(f"IBKR Python override: {args.ibkr_python}")

    if args.skip_schwab and args.skip_ibkr:
        raise ValueError("At least one source must be enabled (Schwab or IBKR).")

    print("\n[1/5] Preflight checks...")
    if not args.skip_schwab:
        check_schwab_token()
    if args.refresh_schwab and not args.skip_schwab:
        require_module(
            module_name="schwabdev",
            source_label="Schwab",
            pip_hint="pip install schwabdev",
            python_executable=schwab_python,
        )
    if args.refresh_ibkr and not args.skip_ibkr:
        require_module(
            module_name="ib_insync",
            source_label="IBKR",
            pip_hint="pip install ib_insync nest_asyncio",
            python_executable=ibkr_python,
        )
        require_module(
            module_name="nest_asyncio",
            source_label="IBKR",
            pip_hint="pip install ib_insync nest_asyncio",
            python_executable=ibkr_python,
        )
    if not args.skip_ibkr and args.refresh_ibkr:
        ensure_ibkr_ready(
            port=args.ibkr_port,
            tws_app=args.tws_app,
            wait_seconds=args.tws_wait_seconds,
        )

    print("\n[2/5] Source refresh...")
    if args.refresh_schwab and not args.skip_schwab:
        run_source_script(schwab_script, "Schwab", schwab_python)
    if args.refresh_ibkr and not args.skip_ibkr:
        run_source_script(ibkr_script, "IBKR", ibkr_python)

    print("\n[3/5] Loading source files...")
    source_frames = []

    if not args.skip_schwab:
        schwab_raw = load_excel(schwab_file, "Schwab")
        schwab_norm = normalize_schwab(schwab_raw, include_cash=args.include_cash)
        source_frames.append(schwab_norm)
        print_source_summary(schwab_norm, "Schwab (normalized)")

    if not args.skip_ibkr:
        ibkr_raw = load_excel(ibkr_file, "IBKR")
        ibkr_norm = normalize_ibkr(ibkr_raw, include_cash=args.include_cash)
        source_frames.append(ibkr_norm)
        print_source_summary(ibkr_norm, "IBKR (normalized)")

    if not source_frames:
        raise RuntimeError("No holdings loaded. Check source paths and skip flags.")

    print("\n[4/5] Aggregating by symbol...")
    detail_df = pd.concat(source_frames, ignore_index=True)
    aggregated_df = aggregate_positions(detail_df)
    client_df = to_client_schema(aggregated_df)
    print(f"  Combined source rows: {len(detail_df):,}")
    print(f"  Final unique symbols: {len(client_df):,}")

    print("\n[5/5] Writing Client workbook...")
    backup_path = write_client_workbook(
        client_df=client_df,
        output_path=output_path,
        backup=not args.no_backup,
    )

    print("\n" + "=" * 80)
    print("PHASE 0 COMPLETE")
    print("=" * 80)
    print(f"Rows written: {len(client_df):,}")
    print(f"Client file: {output_path}")
    if backup_path:
        print(f"Backup: {backup_path}")


if __name__ == "__main__":
    main()
