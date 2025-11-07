"""
=============================================================================
GS BASKET DATA WITH HEADINGS - Enhanced Bloomberg Integration
=============================================================================

INPUT FILES:
- None (data is fetched directly from Goldman Sachs Marquee API)
- Requires: Active GsSession with valid authentication credentials
- API Credentials: client_id and client_secret (embedded in source)

OUTPUT FILES:
- /Users/macbook2024/Library/CloudStorage/Dropbox/AAA Backup/A Working/News/Data Collection/GSCB_FLAGSHIP_coverage_with_desc_ALL.xlsx
  Description: Enhanced Goldman Sachs basket data with specific column headings and BDP formulas for Bloomberg integration
  Format: Excel file with custom column structure and embedded Bloomberg formulas
  Contents: Ticker symbols, asset IDs, names, descriptions, and BDP formulas for real-time data retrieval

VERSION HISTORY:
v1.0.0 (2025-10-10): Initial release with basic data retrieval
v1.1.0 (2025-10-17): Added custom column headings and BDP formula integration
v1.2.0 (2025-11-06): Enhanced documentation and Bloomberg optimization

DEPENDENCIES:
- gs_quant (Goldman Sachs API library)
- pandas
- openpyxl (for Excel file manipulation and formula insertion)

AUTHOR: Generated from Jupyter Notebook

DESCRIPTION:
This script retrieves Goldman Sachs basket coverage data from the Marquee API
and creates an output file with the specific column headings requested:
Index Name, Ticker, Bloomberg, CHG_PCT_1D, CURRENT_TRR_1WK, CHG_PCT_YTD, 
CHG_PCT_1YR, CHG_PCT_3YR, LONG_COMP_NAME, REGION_OR_COUNTRY, SECURITY_TYPI

The enhanced version includes BDP (Bloomberg Data Point) formulas for real-time
data integration with Bloomberg terminals.

KEY FEATURES:
1. API Integration: Connects to Goldman Sachs Marquee API for data retrieval
2. Custom Column Headings: Implements specific column structure for Bloomberg compatibility
3. BDP Formula Integration: Adds Bloomberg formulas for real-time data retrieval
4. Batch Processing: Processes data in batches of 200 to comply with API rate limits
5. Asset Resolution: Automatically resolves ticker symbols to asset IDs
6. Metadata Enrichment: Fetches detailed descriptions and metadata for each basket
7. Bloomberg Integration: Populates Bloomberg column and adds BDP formulas to columns H-L

DATA PROCESSING WORKFLOW:
1. Connect to GSCB_FLAGSHIP dataset via Marquee API
2. Retrieve coverage data with basic identifiers
3. Build unique ticker list maintaining original order
4. Create ticker-to-asset-ID mapping from coverage data
5. Resolve missing asset IDs via API calls
6. Fetch full asset metadata including descriptions
7. Convert asset objects to structured DataFrame with requested headings
8. Merge metadata with resolved tickers
9. Populate Bloomberg column (ticker + " Index")
10. Export enriched data to Excel file
11. Load workbook and add BDP formulas to columns H-L
12. Save workbook with embedded Bloomberg formulas

BLOOMBERG INTEGRATION:
- Column G: Bloomberg (ticker + " Index")
- Columns H-L: BDP formulas for real-time data
  - H: CHG_PCT_1D (=BDP($G2,H$1))
  - I: CURRENT_TRR_1WK (=BDP($G2,I$1))
  - J: CHG_PCT_YTD (=BDP($G2,J$1))
  - K: CHG_PCT_1YR (=BDP($G2,K$1))
  - L: CHG_PCT_3YR (=BDP($G2,L$1))

USAGE:
python gs_basket_data_with_headings.py

REQUIREMENTS:
- Valid Goldman Sachs API credentials (embedded in source)
- Internet connection for API access
- gs_quant library installation
- openpyxl for Excel file manipulation
- Bloomberg terminal access for BDP formula functionality

NOTES:
- The GSCB_FLAGSHIP dataset contains Goldman's flagship basket products
- Data is fetched in batches of 200 to comply with API rate limits
- Output includes the requested column headings with available data
- BDP formulas enable real-time data retrieval in Bloomberg environment
- Runtime varies based on API response times and dataset size
"""

try:
    import pandas as pd
    from gs_quant.data import Dataset
    from gs_quant.session import GsSession
    from gs_quant.api.gs.assets import GsAssetApi
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌ Required library not found: {e}")
    print("Please install the required dependencies:")
    print("pip install gs-quant pandas openpyxl")
    print("or")
    print("pip install gs_quant pandas openpyxl")
    exit(1)

# Authentication credentials
client_id = '6d7fb0d257c44d81a3cf25690a795aa0'
client_secret = 'cf4dd1ad1b0a575bd44156763bf7d920ffcd11d766621cfe61bd155aa518d343'

# Connect to Goldman Sachs session with read_product_data scope
GsSession.use(client_id=client_id, client_secret=client_secret, scopes=('read_product_data',))

# Output file path
OUTPUT_FILE = 'GSCB_FLAGSHIP_coverage_with_desc_ALL.xlsx'

def chunks(seq, n=200):
    """Split a sequence into chunks of size n"""
    for i in range(0, len(seq), n):
        yield seq[i:i+n]

# STEP 1: CONNECT TO DATASET
# Get the dataset - GSCB_FLAGSHIP contains Goldman Sachs flagship basket products
basket_ds = Dataset('GSCB_FLAGSHIP')

# STEP 2: RETRIEVE COVERAGE DATA
# Get the coverage - returns all available baskets with basic identifiers
print("Fetching basket coverage data from Goldman Sachs API...")
coverage = basket_ds.get_coverage()

# STEP 3: CONVERT TO DATAFRAME
# Convert the coverage to a DataFrame for easier manipulation
df = pd.DataFrame(coverage)
print(f"Retrieved {len(df)} baskets from coverage data")

# STEP 4: IDENTIFY COLUMNS
# Automatically detect which column names are used for ticker and asset ID
ticker_col = next((c for c in ['ticker','bbid','Ticker','BBID','TICKER'] if c in df.columns), None)
bbid_col = next((c for c in ['bbid', 'Bloomberg'] if c in df.columns), None)
id_col = next((c for c in ['assetId','AssetId','id','ID','marqueeId','mqid'] if c in df.columns), None)

# STEP 5: BUILD UNIQUE TICKER LIST
# Create a list of unique tickers while maintaining the original order
if ticker_col:
    tickers_all = df[ticker_col].dropna().astype(str).str.strip().tolist()
    tickers_unique = list(dict.fromkeys(tickers_all))  # preserves order
    print(f"Found {len(tickers_unique)} unique tickers")
else:
    tickers_unique = []

# STEP 6: BUILD TICKER-TO-ID MAPPING FROM COVERAGE DATA
# Create mapping of ticker symbols to asset IDs from the coverage data
id_map = {}
if ticker_col and id_col:
    tmp = df[[ticker_col, id_col]].dropna().astype(str)
    # if duplicates, keep the first occurrence to preserve order
    seen = set()
    for t, i in zip(tmp[ticker_col], tmp[id_col]):
        if t not in seen:
            id_map[t] = i
            seen.add(t)

# STEP 7: RESOLVE MISSING ASSET IDs VIA API
# For any tickers that don't have an asset ID in the coverage data, query the API
if ticker_col:
    to_resolve = [t for t in tickers_unique if t not in id_map]
    print(f"Resolving {len(to_resolve)} additional tickers via API...")
    for batch in chunks(to_resolve, 200):
        # returns mapping: queryTicker -> [ {id: ..., ticker: ...}, ... ]
        resp = GsAssetApi.resolve_assets(identifier=batch, fields=['id','ticker'], limit=1)
        for q, hits in resp.items():
            if hits and 'id' in hits[0]:
                id_map[q] = hits[0]['id']

# STEP 8: HANDLE CASES WITH ONLY ASSET IDs (NO TICKERS)
# If the coverage data has asset IDs but no ticker column, create placeholder tickers
if not ticker_col and id_col:
    ids_in_file = df[id_col].dropna().astype(str).tolist()
    ids_unique = list(dict.fromkeys(ids_in_file))
    # we will still fetch meta for these IDs; set a placeholder ticker for join
    tickers_unique = [f'__row_{i:06d}__' for i in range(len(ids_unique))]
    id_map = dict(zip(tickers_unique, ids_unique))  # placeholder ticker -> id
    print(f"Processing {len(ids_unique)} asset IDs without ticker information")

# STEP 9: CREATE RESOLVED DATAFRAME
# Build a DataFrame with tickers and their resolved asset IDs in original order
resolved = pd.DataFrame({'ticker': tickers_unique,
                         'AssetId': [id_map.get(t) for t in tickers_unique]})

# STEP 10: FETCH FULL ASSET METADATA FROM API
# Query the Marquee API to get complete asset information (name, description)
# for all resolved asset IDs
print("Fetching detailed asset metadata from API...")
ids = [i for i in resolved['AssetId'] if pd.notna(i)]
assets = []
for batch in chunks(ids, 200):
    assets.extend(GsAssetApi.get_many_assets(id=batch))  # returns GsAsset objects

# STEP 11: CONVERT ASSET OBJECTS TO DATAFRAME WITH REQUESTED HEADINGS
# Create DataFrame with all requested column headings
meta = pd.DataFrame([{
    'AssetId': getattr(a, 'id', None),
    'name': getattr(a, 'name', None),
    'description': getattr(a, 'description', None),
    # Existing fields
    'api_ticker': getattr(a, 'ticker', None) or getattr(a, 'bbid', None),
    # New column headings with available data (empty values for financial metrics)
    'Index Name': getattr(a, 'name', None),
    'Ticker': getattr(a, 'ticker', None) or getattr(a, 'bbid', None),
    'Bloomberg': None,
    'CHG_PCT_1D': None,
    'CURRENT_TRR_1WK': None,
    'CHG_PCT_YTD': None,
    'CHG_PCT_1YR': None,
    'CHG_PCT_3YR': None,
    'LONG_COMP_NAME': getattr(a, 'name', None),
    'REGION_OR_COUNTRY': None,
    'SECURITY_TYPI': None,
} for a in assets])

# STEP 12: MERGE METADATA WITH RESOLVED TICKERS
# Combine the resolved ticker/ID pairs with the enriched metadata from API
out = resolved.merge(meta, on='AssetId', how='left')

# STEP 13: PREPARE FINAL OUTPUT WITH ALL REQUESTED COLUMNS
# Select all columns including the new headings
if ticker_col:
    # Use the original ticker column from the coverage data
    final = out[['ticker', 'AssetId', 'name', 'description', 'Index Name', 'Ticker', 'Bloomberg',
                'CHG_PCT_1D', 'CURRENT_TRR_1WK', 'CHG_PCT_YTD', 'CHG_PCT_1YR', 'CHG_PCT_3YR',
                'LONG_COMP_NAME', 'REGION_OR_COUNTRY', 'SECURITY_TYPI']].copy()
else:
    # When no ticker in file, use api_ticker and rename it to ticker
    final = out[['api_ticker', 'AssetId', 'name', 'description', 'Index Name', 'Ticker', 'Bloomberg',
                'CHG_PCT_1D', 'CURRENT_TRR_1WK', 'CHG_PCT_YTD', 'CHG_PCT_1YR', 'CHG_PCT_3YR',
                'LONG_COMP_NAME', 'REGION_OR_COUNTRY', 'SECURITY_TYPI']].rename(columns={'api_ticker': 'ticker'})

# STEP 13B: POPULATE BLOOMBERG COLUMN
# Bloomberg column = ticker (column A) + " Index"
final['Bloomberg'] = final['ticker'] + ' Index'

# STEP 14: WRITE OUTPUT FILE
# Export enriched data to Excel file with all requested column headings
final.to_excel(OUTPUT_FILE, index=False)
print(f"Successfully wrote {len(final)} records to {OUTPUT_FILE} with all requested column headings")

# STEP 15: ADD BDP FORMULAS TO COLUMNS H-L
# Load workbook and add BDP formulas with proper cell references
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

# Map column letters to header row values for BDP formula fields
# Columns H-L correspond to: CHG_PCT_1D, CURRENT_TRR_1WK, CHG_PCT_YTD, CHG_PCT_1YR, CHG_PCT_3YR
bdp_columns = ['H', 'I', 'J', 'K', 'L']
bloomberg_col = 'G'  # Bloomberg column with ticker + " Index"

# Add BDP formulas starting from row 2 (row 1 is header)
for row in range(2, len(final) + 2):
    for col_letter in bdp_columns:
        # Formula: =BDP($G2,H$1) where G is Bloomberg col, and H$1 is the header
        formula = f'=BDP(${bloomberg_col}${row},{col_letter}$1)'
        ws[f'{col_letter}{row}'] = formula

# Save workbook with formulas
wb.save(OUTPUT_FILE)
print(f"Added BDP formulas to columns H-L (CHG_PCT_1D through CHG_PCT_3YR)")

# STEP 16: DIAGNOSTICS AND REPORTING
# Report any tickers that couldn't be resolved
unresolved = [t for t in tickers_unique if id_map.get(t) is None]
if unresolved:
    print(f'Warning: Unresolved tickers (no AssetId found): {len(unresolved)}')

missing_desc = final[final['description'].isna()]['ticker'].tolist()
if missing_desc:
    print(f'Warning: No description populated for {len(missing_desc)} tickers')
