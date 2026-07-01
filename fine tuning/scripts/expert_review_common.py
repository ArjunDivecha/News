#!/usr/bin/env python3
"""
Shared helpers for the expert-review ETF classifier training round.
"""

from __future__ import annotations

import json
import math
import random
import re
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[2]
FINE_TUNING_ROOT = REPO_ROOT / "fine tuning"
DEFAULT_UNIVERSE_PATH = REPO_ROOT / "data" / "universe.xlsx"
DEFAULT_OUTPUT_DIR = FINE_TUNING_ROOT / "outputs" / "expert_review"

TIER1_VALUES = [
    "Equities",
    "Fixed Income",
    "Commodities",
    "Currencies (FX)",
    "Multi-Asset / Thematic",
    "Volatility / Risk Premia",
    "Alternative / Synthetic",
    "Factor",
]

VERDICT_VALUES = ["llama", "deepseek", "both_wrong", "both_right"]

ALLOWED_TIER3_TAGS = {
    "Equity",
    "Credit",
    "FX",
    "Commodity",
    "Multi-Asset",
    "Volatility",
    "Alternative",
    "US",
    "Europe",
    "Asia",
    "EM",
    "Global",
    "China",
    "Japan",
    "India",
    "Canada",
    "UAE",
    "APAC",
    "Australia",
    "Developed",
    "Tech",
    "Energy",
    "Financials",
    "Healthcare",
    "Industrials",
    "Consumer",
    "Defensive",
    "ESG",
    "Dividend",
    "Growth",
    "Value",
    "Momentum",
    "Quality",
    "Infrastructure",
    "Real Estate",
    "Utilities",
    "Materials",
    "Active",
    "Passive",
    "Equal-Weight",
    "Thematic",
    "Quantitative",
    "Options-Based",
    "Factor-Based",
    "Low Volatility",
    "Domestic",
    "International",
    "Diversified",
    "Investment Grade",
    "High Yield",
    "Short Duration",
    "Medium Duration",
    "Long Duration",
    "IG Credit",
    "HY Credit",
    # --- 2026-07 additions (holdings coverage gaps) ---
    "Treasury",       # sovereign/government bonds (e.g. TLT) - distinct from Credit
    "Municipal",      # muni bond funds (e.g. VCLAX)
    "Small-Cap",
    "Mid-Cap",
    "Large-Cap",
}

TAG_ALIASES = {
    "Commodities": "Commodity",
    "Taiwan": "Asia",
    # --- 2026-07 normalization: fold universe/model variants to canonical ---
    "Precious Metals": "Materials",
    "Metals": "Materials",
    "Gold": "Materials",
    "Short (<2Y)": "Short Duration",
    "Medium (2-10Y)": "Medium Duration",
    "Long (>10Y)": "Long Duration",
    "Large-cap": "Large-Cap",
    "Large Cap": "Large-Cap",
    "Small Cap": "Small-Cap",
    "Small-cap": "Small-Cap",
    "Mid-cap": "Mid-Cap",
    "Mid Cap": "Mid-Cap",
    "Government": "Treasury",
    "Sovereign": "Treasury",
    "Rates": "Treasury",
    "Muni": "Municipal",
    "Munis": "Municipal",
}

SYSTEM_PROMPT = """You are an expert financial asset classification specialist. Classify the given asset into a structured taxonomy.

TIER-1 CATEGORIES (pick exactly one):
- Equities: Stock indices, equity ETFs, equity-focused baskets, equity indices
- Fixed Income: Bonds, credit, yield-focused instruments, fixed income ETFs
- Commodities: Energy, metals, agriculture, commodity indices
- Currencies (FX): Currency pairs and FX instruments
- Multi-Asset / Thematic: Cross-asset, thematic baskets, macro themes, multi-asset indices
- Volatility / Risk Premia: VIX, volatility indices, carry strategies, risk premia
- Alternative / Synthetic: Quantitative baskets, factor portfolios, proprietary constructs, custom indices
- Factor: Factor ETFs used as report factor proxies

TIER-2 CATEGORIES (examples by Tier-1):
- Equities: Global Indices | Sector Indices | Country/Regional | Thematic/Factor | Real Estate / REITs
- Fixed Income: Sovereign Bonds | Corporate Credit | Credit Spreads | Yield Curves | Broad Fixed Income
- Commodities: Energy | Metals | Agriculture | Broad Commodities
- Currencies: Majors (EUR/USD, GBP/USD, USD/JPY) | EM FX | Broad Currency
- Multi-Asset: Cross-Asset Indices | Inflation/Growth Themes | Thematic Baskets
- Volatility: Vol Indices | Carry/Value Factors | Risk Premia Strategies
- Alternative: Quant/Style Baskets | Custom/Proprietary | Factor-Based
- Factor: Equity Factor | Rates Factor | Credit Factor | Commodity Factor | FX Factor | Volatility Factor

TIER-3 TAGS (select all that apply from):
- Asset Class: Equity | Credit | Treasury | Municipal | FX | Commodity | Multi-Asset | Volatility | Alternative
- Region: US | Europe | Asia | EM | Global | China | Japan | India | Canada | UAE | APAC | Australia | Developed
- Sector/Theme: Tech | Energy | Financials | Healthcare | Industrials | Consumer | Defensive | ESG | Dividend | Growth | Value | Momentum | Quality | Infrastructure | Real Estate | Utilities | Materials
- Strategy: Active | Passive | Equal-Weight | Thematic | Quantitative | Options-Based | Dividend | Factor-Based | Low Volatility | Defensive | Domestic | International | Diversified
- Size: Small-Cap | Mid-Cap | Large-Cap
- Duration/Credit: Investment Grade | High Yield | Short Duration | Medium Duration | Long Duration | IG Credit | HY Credit

CLASSIFICATION GUIDANCE:
- Treasury = sovereign/government bonds (US Treasuries, gilts); Credit = corporate/credit bonds; Municipal = muni bonds. A long-dated US Treasury fund is "Treasury, Long Duration", NOT "Credit".
- A long/short, market-neutral, or relative-value fund (e.g. long value / short growth) is "Alternative", even if its underlying names are equities. Tag its style too (e.g. "Alternative, Value").
- Apply "Passive" to index-tracking funds and "Active" to actively-managed funds. Do NOT put Active or Passive on a single company's stock — an individual stock takes only its Region/Sector/Size/Style tags.
- Size tags: apply AT MOST ONE (Small-Cap OR Mid-Cap OR Large-Cap), and ONLY when size is a defining tilt (e.g. a dedicated small-cap fund). For a broad or all-cap fund (most single-country and total-market funds), omit size entirely — do NOT list multiple size tags.
- Every asset must get at least an Asset Class tag and, for regional/country funds, a Region tag. A multi-asset/allocation fund is "Multi-Asset"; a leveraged/inverse or market-neutral product is "Alternative"; a municipal-bond fund is "Municipal". Never return an empty tag list.

RESPOND ONLY with a JSON object in this exact format:
{"ticker": "TICKER", "tier1": "Category", "tier2": "Sub-category", "tier3_tags": ["tag1", "tag2", "tag3"]}

No explanation, no markdown formatting, just the JSON."""


def clean_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_label(value: Any) -> str:
    return clean_text(value)


def parse_tags(value: Any) -> list[str]:
    if isinstance(value, list):
        raw = value
    elif isinstance(value, tuple) or isinstance(value, set):
        raw = list(value)
    else:
        if value is None or pd.isna(value):
            return []
        text = str(value).strip()
        if not text or text.lower() in {"nan", "none", "null"}:
            return []
        try:
            parsed = json.loads(text)
            raw = parsed if isinstance(parsed, list) else [text]
        except json.JSONDecodeError:
            raw = text.split(",")
    seen: set[str] = set()
    tags: list[str] = []
    for item in raw:
        tag = clean_text(item)
        if tag and tag.lower() not in seen:
            seen.add(tag.lower())
            tags.append(tag)
    return tags


def tags_to_string(tags: Any) -> str:
    normalized: list[str] = []
    seen: set[str] = set()
    for tag in parse_tags(tags):
        tag = TAG_ALIASES.get(tag, tag)
        if tag in ALLOWED_TIER3_TAGS and tag.lower() not in seen:
            normalized.append(tag)
            seen.add(tag.lower())
    return ", ".join(normalized)


def parse_model_json(response_text: str) -> dict[str, Any]:
    text = response_text.strip()
    if "```json" in text:
        text = text.split("```json", 1)[1].split("```", 1)[0].strip()
    elif "```" in text:
        text = text.split("```", 1)[1].split("```", 1)[0].strip()
    data = json.loads(text)
    return {
        "tier1": data.get("tier1") or data.get("tier_1") or "Unknown",
        "tier2": data.get("tier2") or data.get("tier_2") or "Unknown",
        "tier3_tags": parse_tags(data.get("tier3_tags") or data.get("tier_3") or []),
        "raw_response": response_text,
    }


def universe_to_model_input(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame()
    out["Ticker"] = df["yf_ticker"].map(clean_text)
    out["Name"] = df["name"].map(clean_text)
    out["Bloomberg"] = df.get("bloomberg_ticker", "").map(clean_text)
    out["CIE_DES"] = df["description"].map(clean_text)
    out["FUND_OBJECTIVE_LONG"] = df["description"].map(clean_text)
    out["FUND_ASSET_CLASS_FOCUS"] = df.get("tier1", "").map(clean_text)
    out["FUND_GEO_FOCUS"] = ""
    out["FUND_STRATEGY"] = df.get("tags", "").map(clean_text)
    out["STYLE_ANALYSIS_REGION_FOCUS"] = ""
    return out


def build_user_prompt(row: pd.Series) -> str:
    fields = [
        ("Ticker", row.get("yf_ticker")),
        ("Name", row.get("name")),
        ("Description", row.get("description")),
        ("Existing Tier 1", row.get("tier1")),
        ("Existing Tier 2", row.get("tier2")),
        ("Existing Tags", row.get("tags")),
        ("Bloomberg Ticker", row.get("bloomberg_ticker")),
        ("Source", row.get("source")),
    ]
    lines = ["Classify this ETF:"]
    for label, value in fields:
        text = clean_text(value)
        if text:
            lines.append(f"{label}: {text}")
    return "\n".join(lines)


def demo_predictions(universe: pd.DataFrame) -> pd.DataFrame:
    rows = []
    tier1_cycle = ["Equities", "Fixed Income", "Commodities", "Multi-Asset / Thematic"]
    for idx, row in universe.reset_index(drop=True).iterrows():
        base_tier1 = normalize_label(row.get("tier1")) or "Equities"
        base_tier2 = normalize_label(row.get("tier2")) or "Unknown"
        base_tags = tags_to_string(row.get("tags"))

        llama_tier1 = base_tier1
        llama_tier2 = base_tier2
        deepseek_tier1 = base_tier1
        deepseek_tier2 = base_tier2

        if idx % 31 == 0:
            deepseek_tier1 = tier1_cycle[(tier1_cycle.index(base_tier1) + 1) % len(tier1_cycle)] if base_tier1 in tier1_cycle else "Equities"
        if idx % 17 == 0:
            deepseek_tier2 = "Thematic/Factor" if base_tier2 != "Thematic/Factor" else "Global Indices"

        rows.append(
            {
                "yf_ticker": clean_text(row.get("yf_ticker")),
                "llama_tier1": llama_tier1,
                "llama_tier2": llama_tier2,
                "llama_tier3_tags": base_tags,
                "deepseek_tier1": deepseek_tier1,
                "deepseek_tier2": deepseek_tier2,
                "deepseek_tier3_tags": base_tags,
                "llama_raw_response": "",
                "deepseek_raw_response": "",
            }
        )
    return pd.DataFrame(rows)


def stratified_sample(df: pd.DataFrame, tier_col: str, target_n: int, min_per_group: int, seed: int) -> pd.DataFrame:
    if df.empty or target_n <= 0:
        return df.head(0).copy()

    rng = random.Random(seed)
    groups = [(name, group.copy()) for name, group in df.groupby(tier_col, dropna=False)]
    picks: list[int] = []

    for _, group in groups:
        group_indices = list(group.index)
        rng.shuffle(group_indices)
        take = min(min_per_group, len(group_indices))
        picks.extend(group_indices[:take])

    remaining_target = max(0, min(target_n, len(df)) - len(set(picks)))
    if remaining_target:
        remaining = [idx for idx in df.index if idx not in set(picks)]
        weights = df.loc[remaining, tier_col].map(df[tier_col].value_counts()).rsub(len(df) + 1)
        sampled = pd.DataFrame({"idx": remaining, "weight": weights.to_numpy()})
        sampled = sampled.sample(
            n=min(remaining_target, len(sampled)),
            weights="weight",
            random_state=seed,
            replace=False,
        )
        picks.extend(sampled["idx"].tolist())

    rng.shuffle(picks)
    return df.loc[picks[: min(target_n, len(df))]].copy()


def build_training_example(row: pd.Series, tier1: str, tier2: str, tags: Any) -> dict[str, Any]:
    user_content = build_user_prompt(row)
    assistant_content = json.dumps(
        {
            "ticker": clean_text(row.get("yf_ticker")),
            "tier1": clean_text(tier1),
            "tier2": clean_text(tier2),
            "tier3_tags": parse_tags(tags),
        },
        ensure_ascii=False,
    )
    return {
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_content},
            {"role": "assistant", "content": assistant_content},
        ]
    }


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
            count += 1
    return count


def add_review_workbook_formatting(path: Path, adjudicate_rows: int, tier3_rows: int) -> None:
    wb = load_workbook(path)
    if "adjudicate" in wb.sheetnames:
        ws = wb["adjudicate"]
        ws.freeze_panes = "A2"
        widths = {
            "A": 13,
            "B": 38,
            "C": 55,
            "D": 16,
            "E": 24,
            "F": 24,
            "G": 24,
            "H": 24,
            "I": 24,
            "J": 28,
            "K": 34,
            "L": 16,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        tier1_dv = DataValidation(type="list", formula1=f'"{",".join(TIER1_VALUES)}"', allow_blank=True)
        verdict_dv = DataValidation(type="list", formula1=f'"{",".join(VERDICT_VALUES)}"', allow_blank=True)
        ws.add_data_validation(tier1_dv)
        ws.add_data_validation(verdict_dv)
        if adjudicate_rows:
            tier1_dv.add(f"I2:I{adjudicate_rows + 1}")
            verdict_dv.add(f"L2:L{adjudicate_rows + 1}")

    if "tier3_spotcheck" in wb.sheetnames:
        ws = wb["tier3_spotcheck"]
        ws.freeze_panes = "A2"
        for col, width in {"A": 13, "B": 38, "C": 55, "D": 38, "E": 38, "F": 48}.items():
            ws.column_dimensions[col].width = width

    wb.save(path)


def micro_f1(predicted: Iterable[set[str]], actual: Iterable[set[str]]) -> dict[str, float]:
    tp = fp = fn = 0
    for pred_set, actual_set in zip(predicted, actual):
        pred_lower = {x.lower() for x in pred_set}
        actual_lower = {x.lower() for x in actual_set}
        tp += len(pred_lower & actual_lower)
        fp += len(pred_lower - actual_lower)
        fn += len(actual_lower - pred_lower)
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
    return {"precision": precision, "recall": recall, "f1": f1, "tp": tp, "fp": fp, "fn": fn}


def pct(numerator: float, denominator: float) -> float:
    if denominator == 0 or math.isnan(denominator):
        return 0.0
    return numerator / denominator
